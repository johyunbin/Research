# -*- coding: utf-8 -*-
# Phase 5 통합: 검교정 환경소음망(noiseinfo) 검증 → 새 Fig 10.
#  (a) 2020→2023 추세: S-DoT(하락) vs 검교정 자동(도로)·수동(일반/도로)(안정·상승) = 드리프트 아티팩트 확정.
#  (b) offset: 검교정 LAeq vs 인근 S-DoT(~12 dB 저측정) = 절대레벨 사용 불가.
import os, csv, math, glob, warnings
import statistics as st
from collections import defaultdict
import openpyxl
import matplotlib.pyplot as plt
import figstyle as FS
warnings.filterwarnings("ignore")

FS.apply_style()
ROOT = r"T:\00_BACKUP\01_한양대학교\00_연구실\00_프로젝트\★_논문화 프로젝트\31_환경소음_이동성"
P = os.path.join(ROOT, "data", "processed"); REF = os.path.join(ROOT, "data", "reference"); FIG = os.path.join(ROOT, "_claude", "figures")
def emean(vs):
    vs = [v for v in vs if v is not None]
    return 10*math.log10(sum(10**(v/10) for v in vs)/len(vs)) if vs else None
YRS = ["2020", "2021", "2022", "2023"]

# --- 검교정 자동(일별, 도로 9점) 연추세 ---
wb = openpyxl.load_workbook(os.path.join(REF, "측정자료_환경소음(자동_일별)_데이터조회.xlsx"), data_only=True)
rows = list(wb.active.iter_rows(values_only=True)); h = [str(x) for x in rows[0]]
iL, iD = h.index("LEQ"), h.index("측정일")
auto = defaultdict(list)
for r in rows[1:]:
    try:
        d = str(r[iD]); y = d[:4]; v = float(r[iL])
    except: continue
    if y in YRS and 20 <= v <= 110: auto[y].append(v)
wb.close()
auto_y = {y: emean(auto[y]) for y in YRS}

# --- 검교정 수동(분기, 일반/도로) 연추세 ---
man = defaultdict(lambda: defaultdict(list))
for f in glob.glob(os.path.join(REF, "측정자료_환경소음(수동)_데이터조회*.xlsx")):
    wb = openpyxl.load_workbook(f, data_only=True); rr = list(wb.active.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rr) if r and "번호" in [str(x) for x in r]); hh = [str(x) for x in rr[hi]]
    iz, iy = hh.index("지역"), hh.index("측정연도"); iavg = [i for i, x in enumerate(hh) if x == "평균"][0]
    for r in rr[hi+1:]:
        if not r or r[0] is None: continue
        try: y = str(int(float(r[iy]))); v = float(r[iavg])
        except: continue
        if y in YRS: man[r[iz]][y].append(v)
    wb.close()
man_gen = {y: emean(man["일반"][y]) for y in YRS}
man_road = {y: emean(man["도로"][y]) for y in YRS}

# --- S-DoT balanced 연추세(§3.5 정본, 842 센서) ---
sdot_y = {"2020": 49.4, "2021": 47.6, "2022": 47.2, "2023": 47.1}

print("=== 연추세(2020→2023, Δ) ===")
series = [("S-DoT (this network)", sdot_y, FS.ACCENT["noise"], "-o"),
          ("Calibrated automatic, road (n=9, daily)", auto_y, FS.ACCENT["mobility"], "-s"),
          ("Calibrated manual, general (n=91)", "#5b7fa6", None, None)]
for name, d, *_ in [("S-DoT", sdot_y), ("Auto road", auto_y), ("Manual general", man_gen), ("Manual road", man_road)]:
    print(f"  {name:16s}: " + " ".join(f"{d[y]:.2f}" for y in YRS) + f"  Δ={d['2023']-d['2020']:+.2f} dB")

# --- offset (Tier1 phase5_calibval.csv 재사용) ---
ov = list(csv.DictReader(open(os.path.join(P, "phase5_calibval.csv"), encoding="utf-8-sig")))
offs = [float(r["offset_day"]) for r in ov]
print(f"\noffset(S-DoT−LAeq) 주간: 평균 {st.mean(offs):+.2f} dB (n={len(offs)})")

# ===== 새 Fig 10 =====
fig, ax = plt.subplots(1, 2, figsize=(11.0, 5.2))
yy = [2020, 2021, 2022, 2023]
LINES = [("S-DoT (this network)", sdot_y, FS.ACCENT["noise"], "-o"),
         ("Calibrated auto, road", auto_y, "#3d6b54", "-s"),
         ("Calibrated manual, general", man_gen, "#5b7fa6", "-^"),
         ("Calibrated manual, road", man_road, "#C08A5E", "-D")]
for name, d, c, mk in LINES:
    dd = [d[str(y)] - d["2020"] for y in yy]
    ax[0].plot(yy, dd, mk, color=c, lw=1.9, ms=6, mfc="white", mew=1.6, label=name)
ax[0].axhline(0, color=FS.ACCENT["neutral"], lw=.8, ls="--")
ax[0].set_xticks(yy); ax[0].set_xlabel("Year"); ax[0].set_ylabel("Annual level change from 2020 (dB)")
ax[0].set_title("(a)"); ax[0].legend(loc="lower left", fontsize=7.6); ax[0].set_box_aspect(1)
ax[0].annotate(f"{sdot_y['2023']-sdot_y['2020']:+.1f}", (2023, sdot_y["2023"]-sdot_y["2020"]),
               textcoords="offset points", xytext=(6, -2), fontsize=8.5, color=FS.ACCENT["noise"])
# (b) offset 산점
ZC = {"일반": FS.ACCENT["mobility"], "도로": FS.ACCENT["noise"]}
for z in ("일반", "도로"):
    xs = [float(r["cal_day"]) for r in ov if r["zone"] == z]; ys = [float(r["sdot_day"]) for r in ov if r["zone"] == z]
    ax[1].scatter(xs, ys, s=32, color=ZC[z], alpha=.8, edgecolors="white", linewidths=.5,
                  label={"일반": "Residential/general", "도로": "Roadside"}[z])
ax[1].plot([40, 80], [40, 80], ls=":", color="#bbb", lw=1, zorder=0)
ax[1].set_xlim(40, 80); ax[1].set_ylim(40, 80); ax[1].set_box_aspect(1)
ax[1].set_xlabel("Calibrated L$_{Aeq}$ daytime (dB)"); ax[1].set_ylabel("Nearby S-DoT L$_{day}$ (dB)")
ax[1].set_title("(b)"); ax[1].legend(loc="upper left", fontsize=8)
ax[1].text(.04, .68, f"S-DoT reads {st.mean(offs):+.1f} dB\nvs calibrated (n={len(offs)})",
           transform=ax[1].transAxes, fontsize=8, bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#bbb"))
plt.tight_layout()
out = os.path.join(FIG, "fig_drift_validation.png")   # Fig 10 덮어쓰기(업그레이드)
plt.savefig(out)
print(f"\n-> {out} (새 Fig 10)")

with open(os.path.join(P, "phase5_calibration_trends.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f); w.writerow(["series"] + YRS + ["delta"])
    for name, d in [("sdot", sdot_y), ("cal_auto_road", auto_y), ("cal_manual_general", man_gen), ("cal_manual_road", man_road)]:
        w.writerow([name] + [f"{d[y]:.2f}" for y in YRS] + [f"{d['2023']-d['2020']:+.2f}"])
print("-> phase5_calibration_trends.csv")
