# -*- coding: utf-8 -*-
# 스파이크 공정재분석: 에너지평균 LAeq · 주단위 · 인근3센서 군집 · 주야분리
import os, zipfile, io, csv, math
from collections import defaultdict
from datetime import date
import openpyxl

tmp=r"C:\Users\wh850\AppData\Local\Temp"
ref_zip=os.path.join(tmp,"noise_dl","file_seq4.bin")
sdot_dir=os.path.join(tmp,"sdot")
stations={"시청":(37.56472,126.97694),"신사":(37.51288,127.01116),
          "신촌":(37.55528,126.93694),"성수":(37.548534,127.062747)}
def decode(b):
    for e in ("cp949","euc-kr","utf-8-sig","utf-8"):
        try: return b.decode(e)
        except: pass
    return b.decode("utf-8","replace")
def hav(a,b):
    R=6371000.0; la1,lo1,la2,lo2=map(math.radians,(a[0],a[1],b[0],b[1]))
    h=math.sin((la2-la1)/2)**2+math.cos(la1)*math.cos(la2)*math.sin((lo2-lo1)/2)**2
    return 2*R*math.asin(math.sqrt(h))
def cidx(h,p):
    for i,x in enumerate(h):
        if p in (x or ""): return i
    return -1
def pear(xs,ys):
    n=len(xs)
    if n<8: return None
    mx=sum(xs)/n; my=sum(ys)/n
    sxx=sum((x-mx)**2 for x in xs); syy=sum((y-my)**2 for y in ys); sxy=sum((x-mx)*(y-my) for x,y in zip(xs,ys))
    return sxy/math.sqrt(sxx*syy) if sxx>0 and syy>0 else None
def eavg(vals):
    vv=[v for v in vals if v and v>0]
    return 10*math.log10(sum(10**(v/10) for v in vv)/len(vv)) if vv else None
def align(d1,d2):
    ks=sorted(set(d1)&set(d2)); return [d1[k] for k in ks],[d2[k] for k in ks],ks
def weekly(dmap):
    g=defaultdict(list)
    for ds,v in dmap.items():
        y,m,d=map(int,ds.split("-")); iso=date(y,m,d).isocalendar(); g[(iso[0],iso[1])].append(v)
    return {k:(10*math.log10(sum(10**(x/10) for x in vs)/len(vs))) for k,vs in g.items()}

wb=openpyxl.load_workbook(os.path.join(sdot_dir,"location.xlsx"),read_only=True)
rows=list(wb.active.iter_rows(values_only=True)); hdr=[str(x) if x is not None else "" for x in rows[0]]
lat_i=cidx(hdr,"위도"); lon_i=cidx(hdr,"경도"); ser2c={}
for r in rows[1:]:
    if r[1]:
        try: ser2c[str(r[1]).strip()]=(float(r[lat_i]),float(r[lon_i]))
        except: pass
near={st:[s for _,s in sorted((hav(co,c),s) for s,c in ser2c.items())[:3]] for st,co in stations.items()}
allcand=set(s for v in near.values() for s in v)
print("nearest-3 per station:",{st:[(s,round(hav(stations[st],ser2c[s]))) for s in near[st]] for st in stations})

# reference: daily energy LAeq + official day(낮)/night(밤)
ref_d=defaultdict(dict); ref_day=defaultdict(dict); ref_ngt=defaultdict(dict)
z=zipfile.ZipFile(ref_zip)
for n in z.namelist():
    if n.lower().endswith(".xlsx") and "/2022/" in n.replace("\\","/"):
        wbk=openpyxl.load_workbook(io.BytesIO(z.read(n)),read_only=True,data_only=True)
        for sh in wbk.sheetnames:
            st=sh.split("(")[0].strip()
            if st not in stations: continue
            for row in wbk[sh].iter_rows(values_only=True):
                dd=row[1] if len(row)>1 else None
                if hasattr(dd,"year"):
                    hrs=[row[i] if (i<len(row) and isinstance(row[i],(int,float))) else None for i in range(2,26)]
                    ds=f"{dd.year:04d}-{dd.month:02d}-{dd.day:02d}"
                    la=eavg(hrs)
                    if la is not None: ref_d[st][ds]=la
                    dv=row[26] if len(row)>26 and isinstance(row[26],(int,float)) and row[26]>0 else None
                    nv=row[27] if len(row)>27 and isinstance(row[27],(int,float)) and row[27]>0 else None
                    if dv: ref_day[st][ds]=dv
                    if nv: ref_ngt[st][ds]=nv

# S-DoT: per sensor per date hourly list
sd=defaultdict(lambda:defaultdict(list))   # serial -> date -> [(hour,dB)]
z2=zipfile.ZipFile(os.path.join(sdot_dir,"sdot2022.zip"))
for n in [x for x in z2.namelist() if x.lower().endswith(".csv")]:
    rdr=csv.reader(io.StringIO(decode(z2.read(n)))); h=next(rdr)
    si=cidx(h,"시리얼"); ni=cidx(h,"소음"); ri=cidx(h,"등록")
    for r in rdr:
        if si<len(r) and r[si].strip() in allcand:
            v=r[ni].strip() if ni<len(r) else ""; ts=r[ri].strip() if ri<len(r) else ""
            if not v or not ts: continue
            try: fv=float(v)
            except: continue
            dp=ts.split(); p=dp[0].replace(".","-").split("-")
            if len(p)!=3 or not all(x.isdigit() for x in p): continue
            ds=f"{int(p[0]):04d}-{int(p[1]):02d}-{int(p[2]):02d}"
            hh=None
            if len(dp)>1:
                try: hh=int(dp[1].split(":")[0])
                except: hh=None
            sd[r[si].strip()][ds].append((hh,fv))
print("parsed.")

def sensor_daily(serial, kind="all", minh=18):
    out={}
    for ds,lst in sd.get(serial,{}).items():
        if kind=="all": vals=[v for _,v in lst]
        elif kind=="day": vals=[v for hh,v in lst if hh is not None and 6<=hh<=21]
        else: vals=[v for hh,v in lst if hh is not None and (hh>=22 or hh<=5)]
        if len(vals)>= (minh if kind=="all" else 6):
            e=eavg(vals)
            if e is not None: out[ds]=e
    return out
def cluster_daily(serials, minh=18):
    pool=defaultdict(list)
    for s in serials:
        for ds,lst in sd.get(s,{}).items():
            pool[ds].extend(v for _,v in lst)
    return {ds:eavg(vs) for ds,vs in pool.items() if len(vs)>=minh and eavg(vs) is not None}

print("\n=== FAIR COMPARISON (energy-avg LAeq, 2022) ===")
print("station | nearest single sensor                    | cluster(3 nearest)")
print("        | daily_r  weekly_r  day_r  night_r  bias   | daily_r  weekly_r")
for st in stations:
    s=near[st][0]
    sdl=sensor_daily(s,"all"); x,y,_=align(sdl,ref_d[st]); pr=pear(x,y)
    wk_s=weekly(sdl); wk_r=weekly(ref_d[st]); wx,wy,_=align(wk_s,wk_r); wr=pear(wx,wy)
    sdday=sensor_daily(s,"day"); dx,dy,_=align(sdday,ref_day[st]); dr=pear(dx,dy)
    sdngt=sensor_daily(s,"night"); nx,ny,_=align(sdngt,ref_ngt[st]); nr=pear(nx,ny)
    bias=(sum(b-a for a,b in zip(x,y))/len(x)) if x else float('nan')
    # cluster
    cl=cluster_daily(near[st]); cx,cy,_=align(cl,ref_d[st]); cpr=pear(cx,cy)
    wcl=weekly(cl); wcx,wcy,_=align(wcl,wk_r); cwr=pear(wcx,wcy)
    f=lambda v:(f"{v:.3f}" if v is not None else " na ")
    print(f"{st:>4}    | {f(pr)}  {f(wr)}   {f(dr)}  {f(nr)}  {bias:+5.1f} | {f(cpr)}  {f(cwr)}   (single={s} {hav(stations[st],ser2c[s]):.0f}m)")
print("\n=== DONE ===")
