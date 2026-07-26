# -*- coding: utf-8 -*-
# Phase 2-D 교차검증: 공식 도로교통 4점 LAeq(OA-15473) 연 추세 vs 근접 S-DoT 연 추세.
# 목적: S-DoT 하향 드리프트가 '센서 아티팩트'인지(공식은 안정/상승인데 S-DoT만 하락) 판정.
# + 2022<->2023 스키마 경계 점검(공식 대조). 4점 xlsx = 시트별 일x24시간 LAeq.
import os, io, zipfile, math
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import figstyle as FS

FS.apply_style()
ROOT = r"T:\00_BACKUP\01_한양대학교\00_연구실\00_프로젝트\★_논문화 프로젝트\31_환경소음_이동성"
P = os.path.join(ROOT, "data", "processed")
FIG = os.path.join(ROOT, "_claude", "figures")
ZIP = os.path.join(ROOT, "data", "reference", "seoul_roadtraffic_LAeq_2021-2024.zip")
STATIONS = {"시청": (37.56472, 126.97694), "신사": (37.51288, 127.01116),
            "신촌": (37.55528, 126.93694), "성수": (37.548534, 127.062747)}

def Lmean(vals):
    vs = [v for v in vals if isinstance(v, (int, float)) and v and v > 0]
    return 10 * math.log10(sum(10 ** (v / 10) for v in vs) / len(vs)) if vs else None

# --- 공식 4점 파싱: (station, year) -> 일별 L24 리스트 ---
z = zipfile.ZipFile(ZIP)
xlsxs = [n for n in z.namelist() if n.endswith(".xlsx") and "/202" in n]
off = {st: {} for st in STATIONS}   # st -> year -> [daily L24]
for n in sorted(xlsxs):
    yr = int(n.split("/")[-1][:4])
    if yr > 2023:
        continue
    wb = openpyxl.load_workbook(io.BytesIO(z.read(n)), read_only=True, data_only=True)
    for sh in wb.sheetnames:
        st = next((s for s in STATIONS if s in sh), None)
        if not st:
            continue
        ws = wb[sh]
        for r in ws.iter_rows(min_row=7, values_only=True):
            if r[1] is None:
                continue
            hrs = r[2:26]
            L = Lmean(hrs)
            if L is not None:
                off[st].setdefault(yr, []).append(L)
    wb.close()

print("=== 공식 4점 도로교통 LAeq 연평균 (energy-mean L24) ===")
off_year = {2021: [], 2022: [], 2023: []}
for st in STATIONS:
    vals = {}
    for yr in (2021, 2022, 2023):
        d = off[st].get(yr, [])
        vals[yr] = Lmean(d) if d else None
        if vals[yr]:
            off_year[yr].append(vals[yr])
    tr = (vals[2023] - vals[2021]) if vals[2021] and vals[2023] else float("nan")
    print(f"  {st}: 2021={vals[2021]:.2f}  2022={vals[2022]:.2f}  2023={vals[2023]:.2f}  | 2021->2023 {tr:+.2f} dB")
off_avg = {yr: (sum(v) / len(v)) for yr, v in off_year.items()}
print(f"  [4점 평균] 2021={off_avg[2021]:.2f}  2022={off_avg[2022]:.2f}  2023={off_avg[2023]:.2f}  | 추세 {off_avg[2023]-off_avg[2021]:+.2f} dB")

# --- 근접 S-DoT: 패널 Leq24 연평균 ---
panel = pd.read_csv(os.path.join(P, "analysis_panel.csv"), dtype={"serial": str},
                    usecols=["serial", "lat", "lon", "year", "Leq24"])
coords = panel.groupby("serial")[["lat", "lon"]].first()
def nearest(la, lo):
    d2 = (coords["lat"] - la) ** 2 + (coords["lon"] - lo) ** 2
    return d2.idxmin(), (d2.min() ** 0.5) * 111000  # 대략 m

print("\n=== 근접 S-DoT 연평균 Leq24 (공식과 동일 기간) ===")
sd_year = {2021: [], 2022: [], 2023: []}
for st, (la, lo) in STATIONS.items():
    ser, dist = nearest(la, lo)
    g = panel[panel["serial"] == ser].groupby("year")["Leq24"].mean()
    v = {yr: g.get(yr, float("nan")) for yr in (2021, 2022, 2023)}
    for yr in (2021, 2022, 2023):
        if pd.notna(v[yr]):
            sd_year[yr].append(v[yr])
    tr = v[2023] - v[2021]
    print(f"  {st}(~{dist:.0f}m, {ser}): 2021={v[2021]:.2f}  2022={v[2022]:.2f}  2023={v[2023]:.2f}  | 추세 {tr:+.2f} dB")
sd_avg = {yr: (sum(v) / len(v)) for yr, v in sd_year.items()}
print(f"  [근접4 평균] 2021={sd_avg[2021]:.2f}  2022={sd_avg[2022]:.2f}  2023={sd_avg[2023]:.2f}  | 추세 {sd_avg[2023]-sd_avg[2021]:+.2f} dB")

print("\n=== 판정 ===")
off_tr = off_avg[2023] - off_avg[2021]
sd_tr = sd_avg[2023] - sd_avg[2021]
print(f"공식 4점 추세 {off_tr:+.2f} dB/2y vs 근접 S-DoT 추세 {sd_tr:+.2f} dB/2y")
print(f"괴리(S-DoT - 공식) = {sd_tr - off_tr:+.2f} dB  -> 음(-)이고 크면 S-DoT 하향 드리프트(센서 아티팩트) 입증")

pd.DataFrame([
    ["official_4pt", off_avg[2021], off_avg[2022], off_avg[2023], off_tr],
    ["nearby_sdot", sd_avg[2021], sd_avg[2022], sd_avg[2023], sd_tr],
], columns=["source", "y2021", "y2022", "y2023", "trend_2y"]).to_csv(
    os.path.join(P, "phase2d_drift_crossval.csv"), index=False, encoding="utf-8-sig")
print("-> phase2d_drift_crossval.csv")

# === 드리프트 검증 그림 (robust) ===
yrs4 = panel.groupby("serial")["year"].nunique()
bal = set(yrs4[yrs4 == 4].index)
bsub = panel[panel["serial"].isin(bal)]
ba = bsub.groupby("year")["Leq24"].mean()
bmed = bsub.groupby("year")["Leq24"].median()      # robust(아웃라이어 내성)
city_tr = ba.get(2023) - ba.get(2021)
# 측정소별 추세(2021->2023) — 시청 아웃라이어를 평균에 숨기지 않고 개별 표시
off_tr_st, sd_tr_st = {}, {}
for st, (la, lo) in STATIONS.items():
    ov1, ov3 = Lmean(off[st].get(2021, [])), Lmean(off[st].get(2023, []))
    off_tr_st[st] = (ov3 - ov1) if (ov1 and ov3) else float("nan")
    ser, _ = nearest(la, lo)
    gg = panel[panel["serial"] == ser].groupby("year")["Leq24"].mean()
    sd_tr_st[st] = gg.get(2023, float("nan")) - gg.get(2021, float("nan"))

fig, ax = plt.subplots(1, 2, figsize=(10.5, 5.2))   # 각 패널 정사각형(set_box_aspect)
# (a) balanced 연추세 = 드리프트/다년교란 (mean + median)
yy = [2020, 2021, 2022, 2023]
ax[0].plot(yy, [ba.get(y) for y in yy], "-o", color=FS.ACCENT["noise"], lw=2, ms=7, mfc="white", mew=1.8, label="Mean")
ax[0].plot(yy, [bmed.get(y) for y in yy], "--s", color="#8E9BB5", lw=1.6, ms=6, mfc="white", mew=1.5, label="Median (robust)")
ax[0].set_title("(a)")
ax[0].set_ylabel("Annual L$_{eq,24h}$ (dB)"); ax[0].set_xlabel("Year"); ax[0].set_xticks(yy)
for y in yy:
    ax[0].annotate(f"{ba[y]:.1f}", (y, ba[y]), textcoords="offset points", xytext=(0, 9), ha="center", fontsize=8.5)
ax[0].legend(loc="upper right", fontsize=8.5); ax[0].grid(axis="y", color="#E5E5E0", lw=.7)
# (b) 측정소별 공식 vs S-DoT 추세 산점(1:1선) — 시청 outlier 노출
ax[1].axhline(0, color=FS.ACCENT["neutral"], lw=.6, ls="--"); ax[1].axvline(0, color=FS.ACCENT["neutral"], lw=.6, ls="--")
lim = 8.2
ax[1].plot([-1, lim], [-1, lim], color="#CCCCCC", lw=1, ls=":", zorder=0)
ROM = {"시청": "City Hall", "신사": "Sinsa", "신촌": "Sinchon", "성수": "Seongsu"}
for st in STATIONS:
    x0, y0 = off_tr_st[st], sd_tr_st[st]
    ax[1].scatter(x0, y0, s=70, color=FS.ACCENT["mobility"], edgecolors="white", zorder=3)
    if x0 > lim * 0.7:   # 오른쪽 가장자리 근처면 라벨을 왼쪽으로(축 밖 이탈 방지)
        ax[1].annotate(ROM[st], (x0, y0), textcoords="offset points", xytext=(-8, 4), ha="right", fontsize=9)
    else:
        ax[1].annotate(ROM[st], (x0, y0), textcoords="offset points", xytext=(7, 3), ha="left", fontsize=9)
ax[1].axhspan(-10, 0, color="#E8B8B8", alpha=.12)  # S-DoT 하락 영역(=드리프트 의심)
ax[1].set_xlabel("Official roadside L$_{Aeq}$ trend 2021→2023 (dB)")
ax[1].set_ylabel("Nearby S-DoT trend 2021→2023 (dB)")
ax[1].set_title("(b)")
ax[1].text(.04, .96, f"Citywide balanced S-DoT: {city_tr:+.1f} dB\n(road sites do NOT fall → decline is\nnot uniform hardware drift)",
           transform=ax[1].transAxes, va="top", fontsize=8, bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#bbbbbb"))
ax[1].set_xlim(-1.5, lim); ax[1].set_ylim(-1.5, lim)
for a in ax:
    a.set_box_aspect(1)   # 각 패널 정사각형
plt.tight_layout()
out = os.path.join(FIG, "fig_drift_validation.png")
plt.savefig(out)
print(f"-> {out}  | citywide balanced trend 2021->2023 = {city_tr:+.2f} dB")
print(f"   측정소별 공식추세: {({k: round(v,1) for k,v in off_tr_st.items()})}")
print(f"   측정소별 S-DoT추세: {({k: round(v,1) for k,v in sd_tr_st.items()})}")
