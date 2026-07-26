# -*- coding: utf-8 -*-
# 스파이크 본분석: S-DoT 일평균 dB vs 기준 도로교통 LAeq (2022), 측정소 인근 매칭
import os, zipfile, io, csv, math
from collections import defaultdict
import openpyxl

tmp = r"C:\Users\wh850\AppData\Local\Temp"
ref_zip = os.path.join(tmp, "noise_dl", "file_seq4.bin")
sdot_dir = os.path.join(tmp, "sdot")
stations = {"시청": (37.56472,126.97694), "신사": (37.51288,127.01116),
            "신촌": (37.55528,126.93694), "성수": (37.548534,127.062747)}

def decode(b):
    for e in ("cp949","euc-kr","utf-8-sig","utf-8"):
        try: return b.decode(e)
        except: pass
    return b.decode("utf-8","replace")
def hav(a,b):
    R=6371000.0; la1,lo1,la2,lo2=map(math.radians,(a[0],a[1],b[0],b[1]))
    h=math.sin((la2-la1)/2)**2+math.cos(la1)*math.cos(la2)*math.sin((lo2-lo1)/2)**2
    return 2*R*math.asin(math.sqrt(h))
def cidx(h,p):
    for i,x in enumerate(h):
        if p in (x or ""): return i
    return -1
def pear(xs,ys):
    n=len(xs)
    if n<10: return None
    mx=sum(xs)/n; my=sum(ys)/n
    sxx=sum((x-mx)**2 for x in xs); syy=sum((y-my)**2 for y in ys)
    sxy=sum((x-mx)*(y-my) for x,y in zip(xs,ys))
    if sxx<=0 or syy<=0: return None
    return sxy/math.sqrt(sxx*syy)
def spear(xs,ys):
    def rk(v):
        o=sorted(range(len(v)),key=lambda i:v[i]); r=[0]*len(v)
        for pos,i in enumerate(o): r[i]=pos
        return r
    return pear(rk(xs),rk(ys))
def eavg(vals):
    vv=[v for v in vals if v and v>0]
    return 10*math.log10(sum(10**(v/10) for v in vv)/len(vv)) if vv else None

# S-DoT serial -> coord
wb=openpyxl.load_workbook(os.path.join(sdot_dir,"location.xlsx"),read_only=True)
rows=list(wb.active.iter_rows(values_only=True))
hdr=[str(x) if x is not None else "" for x in rows[0]]
lat_i=cidx(hdr,"위도"); lon_i=cidx(hdr,"경도")
ser2c={}
for r in rows[1:]:
    if r[1]:
        try: ser2c[str(r[1]).strip()]=(float(r[lat_i]),float(r[lon_i]))
        except: pass
cand={st:[s for _,s in sorted((hav(co,c),s) for s,c in ser2c.items())[:6]] for st,co in stations.items()}
allcand=set(s for v in cand.values() for s in v)
dist={(st,s):hav(stations[st],ser2c[s]) for st in stations for s in cand[st]}

# reference 2022 daily LAeq + diurnal
ref_daily=defaultdict(dict); ref_hour=defaultdict(lambda:defaultdict(list))
z=zipfile.ZipFile(ref_zip)
for n in z.namelist():
    if n.lower().endswith(".xlsx") and "/2022/" in n.replace("\\","/"):
        wbk=openpyxl.load_workbook(io.BytesIO(z.read(n)),read_only=True,data_only=True)
        for sh in wbk.sheetnames:
            st=sh.split("(")[0].strip()
            if st not in stations: continue
            for row in wbk[sh].iter_rows(values_only=True):
                d=row[1] if len(row)>1 else None
                if hasattr(d,"year"):
                    hrs=[row[i] if (i<len(row) and isinstance(row[i],(int,float))) else None for i in range(2,26)]
                    ds=f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
                    la=eavg(hrs)
                    if la is not None: ref_daily[st][ds]=la
                    for k,val in enumerate(hrs,start=1):
                        if val and val>0: ref_hour[st][k].append(val)
print("reference 2022 days/station:", {st:len(ref_daily[st]) for st in stations})

# S-DoT 2022 (stream, filter to candidates)
sd_daily=defaultdict(lambda:defaultdict(list)); sd_hour=defaultdict(lambda:defaultdict(list))
z2=zipfile.ZipFile(os.path.join(sdot_dir,"sdot2022.zip"))
for n in [x for x in z2.namelist() if x.lower().endswith(".csv")]:
    rdr=csv.reader(io.StringIO(decode(z2.read(n))))
    h=next(rdr); si=cidx(h,"시리얼"); ni=cidx(h,"소음"); ri=cidx(h,"등록")
    for r in rdr:
        if si<len(r) and r[si].strip() in allcand:
            v=r[ni].strip() if ni<len(r) else ""; ts=r[ri].strip() if ri<len(r) else ""
            if not v or not ts: continue
            try: fv=float(v)
            except: continue
            dp=ts.split(); p=dp[0].replace(".","-").split("-")
            if len(p)!=3 or not all(x.isdigit() for x in p): continue
            ds=f"{int(p[0]):04d}-{int(p[1]):02d}-{int(p[2]):02d}"
            s=r[si].strip(); sd_daily[s][ds].append(fv)
            if len(dp)>1:
                try: sd_hour[s][int(dp[1].split(":")[0])].append(fv)
                except: pass
print("S-DoT 2022 parsed for", len(allcand), "candidate serials.")

out=[("station","serial","dist_m","ndays","pearson","spearman","diurnal_r","bias_ref_minus_sdot","sdot_mean","ref_mean")]
print("\n=== RESULTS: S-DoT daily mean dB vs reference daily LAeq (2022) ===")
for st in stations:
    pick=None
    for s in cand[st]:
        days=[d for d in (set(sd_daily.get(s,{})) & set(ref_daily[st])) if len(sd_daily[s][d])>=18]
        if len(days)>=150: pick=(s,sorted(days)); break
    if not pick:
        bs=None
        for s in cand[st]:
            days=[d for d in (set(sd_daily.get(s,{})) & set(ref_daily[st])) if len(sd_daily[s][d])>=18]
            if bs is None or len(days)>len(bs[1]): bs=(s,sorted(days))
        pick=bs
    if not pick or len(pick[1])<10:
        print(f"[{st}] insufficient overlap"); continue
    s,days=pick
    xs=[sum(sd_daily[s][d])/len(sd_daily[s][d]) for d in days]
    ys=[ref_daily[st][d] for d in days]
    pr=pear(xs,ys); sr=spear(xs,ys); bias=sum(y-x for x,y in zip(xs,ys))/len(xs)
    sdp=[(sum(sd_hour[s][hh])/len(sd_hour[s][hh])) if sd_hour[s].get(hh) else None for hh in range(0,24)]
    rfp=[(sum(ref_hour[st][k])/len(ref_hour[st][k])) if ref_hour[st].get(k) else None for k in range(1,25)]
    pp=[(a,b) for a,b in zip(sdp,rfp) if a is not None and b is not None]
    dr=pear([a for a,_ in pp],[b for _,b in pp]) if len(pp)>=10 else None
    sm=sum(xs)/len(xs); rm=sum(ys)/len(ys)
    drs = ("%.3f"%dr) if dr is not None else "na"
    prs = ("%.3f"%pr) if pr is not None else "na"
    srs = ("%.3f"%sr) if sr is not None else "na"
    print(f"[{st}] S-DoT={s} {dist[(st,s)]:.0f}m days={len(days)} | daily Pearson={prs} Spearman={srs} | diurnal_r={drs} | bias(ref-sdot)={bias:+.1f}dB (sdot {sm:.1f} / ref {rm:.1f})")
    out.append((st,s,f"{dist[(st,s)]:.0f}",len(days),prs,srs,drs,f"{bias:+.1f}",f"{sm:.1f}",f"{rm:.1f}"))

with open(os.path.join(sdot_dir,"spike_summary.csv"),"w",newline="",encoding="utf-8-sig") as f:
    cw=csv.writer(f)
    for row in out: cw.writerow(row)
print("\nsaved spike_summary.csv")
print("=== DONE ===")
