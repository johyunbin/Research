# -*- coding: utf-8 -*-
# 스파이크 점검: (A) 기준 LAeq zip 구조 (B) 4개소 좌표 (C) 인근 S-DoT 매칭거리
import os, zipfile, io, math
import openpyxl

tmp = r"C:\Users\wh850\AppData\Local\Temp"
ref_zip = os.path.join(tmp, "noise_dl", "file_seq4.bin")
sdot_dir = os.path.join(tmp, "sdot")

print("=== (A) reference zip ===")
print("path:", ref_zip, "| exists:", os.path.exists(ref_zip))
if os.path.exists(ref_zip):
    z = zipfile.ZipFile(ref_zip)
    names = z.namelist()
    xlsx = [n for n in names if n.lower().endswith(".xlsx")]
    print("members:", len(names), "| xlsx:", len(xlsx))
    y22 = [n for n in xlsx if "2022" in n]
    print("2022 xlsx:", y22[:14])
    target = None
    for n in y22:
        if "2022-07" in n or "/07" in n or n.split("/")[-1].startswith("2022-07"):
            target = n; break
    if not target and y22: target = y22[0]
    print("opening:", target)
    if target:
        wb = openpyxl.load_workbook(io.BytesIO(z.read(target)), read_only=True, data_only=True)
        print("sheet names:", wb.sheetnames)
        ws = wb[wb.sheetnames[0]]
        print(f"--- sheet '{wb.sheetnames[0]}' first 10 rows (cols<=28) ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10: break
            cells = [round(c,1) if isinstance(c,float) else c for c in row[:28]]
            print("  r%02d:"%i, cells)

print("\n=== (B) station coords (approx) ===")
stations = {
 "시청": (37.56472, 126.97694),
 "신사": (37.51288, 127.01116),
 "신촌": (37.55528, 126.93694),
 "성수": (37.548534, 127.062747),
}
for k,v in stations.items(): print(" ", k, v)

print("\n=== (C) S-DoT serial->coord + nearest ===")
loc = os.path.join(sdot_dir, "location.xlsx")
wb2 = openpyxl.load_workbook(loc, read_only=True)
ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))
hdr = [str(x) if x is not None else "" for x in rows[0]]
def cidx(h, part):
    for i,x in enumerate(h):
        if part in x: return i
    return -1
ser_i = 1
lat_i = cidx(hdr,"위도"); lon_i = cidx(hdr,"경도")
ser2coord = {}
for r in rows[1:]:
    if ser_i < len(r) and r[ser_i]:
        try: ser2coord[str(r[ser_i]).strip()] = (float(r[lat_i]), float(r[lon_i]))
        except: pass
print("serial->coord count:", len(ser2coord))
def hav(a,b):
    R=6371000.0
    la1,lo1,la2,lo2 = map(math.radians,(a[0],a[1],b[0],b[1]))
    h=math.sin((la2-la1)/2)**2+math.cos(la1)*math.cos(la2)*math.sin((lo2-lo1)/2)**2
    return 2*R*math.asin(math.sqrt(h))
for st,co in stations.items():
    near=sorted(((hav(co,c),s) for s,c in ser2coord.items()))[:5]
    print(f"  [{st}] nearest S-DoT:")
    for dist,s in near:
        print(f"     {s}  {dist:6.0f} m")
print("\n=== DONE ===")
